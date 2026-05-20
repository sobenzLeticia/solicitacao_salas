import re
import datetime as dt
from pathlib import Path
from io import BytesIO
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import requests
import base64


# -----------------------  Configurações  -----------------------
# ===============================
# CONFIGURAÇÕES GERAIS
# ===============================

# Caminhos relativos dentro do repositório
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR

CAMINHO_SALAS = DATA_DIR / "SALAS - COPIA.xlsx"
CAMINHO_DISCIPLINAS = DATA_DIR / "Resultados_Gerais.xlsx"
OUTPUT_DIR = BASE_DIR / "resultados"

DIAS_SEMANA = ["SEGUNDA", "TERÇA", "QUARTA", "QUINTA", "SEXTA", "SÁBADO"]
INDICE_DIAS = {d: i for i, d in enumerate(DIAS_SEMANA)}


# ===============================
# FUNÇÕES DE COMMIT NO GITHUB
# ===============================

def commit_dados_disciplinas(df, mensagem=None):
    """
    Salva o DataFrame diretamente no repositório do GitHub via API.
    Substitui o arquivo antigo pelo novo.
    """
    try:
        token = st.secrets["GITHUB_TOKEN"]
        repo = st.secrets["REPO_NAME"]
        branch = st.secrets.get("BRANCH", "main")
    except KeyError as e:
        st.error(f"❌ Secret não configurado: {e}. Vá em Settings → Secrets no Streamlit Cloud.")
        return False
    
    # Nome do arquivo no repositório (mesmo nome que você usa para carregar)
    caminho_arquivo = "Resultados_Gerais.xlsx"
    
    if mensagem is None:
        mensagem = f"Atualiza alocação de salas - {dt.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    # 1. Converte DataFrame para bytes em memória (não salva no disco!)
    buffer = BytesIO()
    df.to_excel(buffer, index=False, engine='openpyxl')
    conteudo_base64 = base64.b64encode(buffer.getvalue()).decode()
    
    # 2. Busca o SHA do arquivo atual (necessário para atualizar)
    url_api = f"https://api.github.com/repos/{repo}/contents/{caminho_arquivo}"
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json"
    }
    
    resp_get = requests.get(url_api, headers=headers, params={"ref": branch})
    
    sha_atual = None
    if resp_get.status_code == 200:
        sha_atual = resp_get.json().get("sha")
    elif resp_get.status_code == 404:
        pass  # Arquivo não existe ainda, vai criar
    else:
        st.error(f"Erro ao buscar arquivo no GitHub: {resp_get.status_code}")
        return False
    
    # 3. Faz o commit (cria ou atualiza)
    payload = {
        "message": mensagem,
        "content": conteudo_base64,
        "branch": branch
    }
    if sha_atual:
        payload["sha"] = sha_atual
    
    resp_put = requests.put(url_api, headers=headers, json=payload)
    
    if resp_put.status_code in [200, 201]:
        return True
    else:
        st.error(f"Erro no commit: {resp_put.status_code} - {resp_put.text}")
        return False


# ===============================
# FUNÇÕES DE LEITURA E PROCESSAMENTO
# ===============================

@st.cache_data(show_spinner=False)
def carregar_dados():
    """Carrega os dados de salas e turmas do repositório."""
    if not CAMINHO_SALAS.exists():
        st.error(f"❌ Arquivo de salas não encontrado em: {CAMINHO_SALAS}")
        st.stop()

    if not CAMINHO_DISCIPLINAS.exists():
        st.error(f"❌ Arquivo de disciplinas não encontrado em: {CAMINHO_DISCIPLINAS}")
        st.stop()

    df_salas = pd.read_excel(CAMINHO_SALAS)
    df_turmas = pd.read_excel(CAMINHO_DISCIPLINAS)
    return df_salas, df_turmas


# -----------------------  Utils horário  -----------------------
def str_to_time(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return None
    if isinstance(s, dt.time):
        return s
    s = str(s).strip()
    for fmt in ("%H:%M:%S", "%H:%M", "%H.%M"):
        try:
            return dt.datetime.strptime(s, fmt).time()
        except Exception:
            pass
    s2 = re.sub(r'[^0-9:]', '', s)
    try:
        return dt.datetime.strptime(s2, "%H:%M").time()
    except Exception:
        return None

def time_to_minutes(t):
    return t.hour * 60 + t.minute

def intervals_overlap(a_start, a_end, b_start, b_end):
    a_s = time_to_minutes(str_to_time(a_start))
    a_e = time_to_minutes(str_to_time(a_end))
    b_s = time_to_minutes(str_to_time(b_start))
    b_e = time_to_minutes(str_to_time(b_end))
    return max(a_s, b_s) < min(a_e, b_e)

# -----------------------  Leitura e processamento  -----------------------
def criar_lista_salas(df_salas: pd.DataFrame):
    salas = []
    for _, row in df_salas.iterrows():
        nome = str(row.get("SALAS") or row.get("SALA") or row.get("NOME") or "").strip()
        capacidade = int(row.get("CAPACIDADE") or 0)
        salas.append({
            "NOME": nome,
            "CAPACIDADE": capacidade,
            "DATAS": set(),
            "HORARIOS_OCUPADOS": set(),
            "HORARIOS_OCUPADOS_SEMANA": {d: [] for d in DIAS_SEMANA},
            "RESERVAS": []
        })
    return salas

def re_split_days(s: str):
    parts = re.split(r'[;,/\\]+|\s{2,}|\s', s)
    return [p for p in parts if p]

def gerar_datas(df_turmas):
    try:
        data_inicio = list(map(int, str(df_turmas.iloc[0, 13]).split(",")))
        data_final = list(map(int, str(df_turmas.iloc[0, 14]).split(",")))
        return pd.date_range(dt.date(*data_inicio), dt.date(*data_final))
    except Exception:
        try:
            col0 = df_turmas.columns[0]
            min_date = pd.to_datetime(df_turmas[col0]).min().date()
            max_date = pd.to_datetime(df_turmas[col0]).max().date()
            return pd.date_range(min_date, max_date)
        except Exception:
            hoje = dt.date.today()
            return pd.date_range(hoje, hoje)

def processar_alocacoes(df_turmas: pd.DataFrame, todas_as_datas, salas_ct: list):
    registros = []
    for _, aloc in df_turmas.iterrows():
        status = str(aloc.get("STATUS") or "").strip()
        if status.upper() != "ALOCADA":
            continue
        sala = str(aloc.get("SALA") or aloc.get("SALAS") or "").strip()
        if not sala:
            continue
        dias_raw = str(aloc.get("DIAS") or "").strip()
        if not dias_raw:
            continue
        dias_tokens = [t.strip().upper() for t in re_split_days(dias_raw)]
        dias_validos = [d for d in dias_tokens if d in INDICE_DIAS]
        if not dias_validos:
            continue
        inicio_raw = aloc.get("HORARIO INICIO") or aloc.get("HORARIO") or aloc.get("HORÁRIO INICIO")
        fim_raw = aloc.get("HORARIO FINAL") or aloc.get("HORÁRIO FINAL") or aloc.get("HORARIO_FIM")
        inicio_t = str_to_time(inicio_raw)
        fim_t = str_to_time(fim_raw)
        descricao = (
            f"{aloc.get('CODIGO') or ''} - "
            f"{aloc.get('DISCIPLINA') or ''} - "
            f"{aloc.get('TURMA') or ''} - "
            f"{aloc.get('PROFESSOR') or ''}"
        )

        indices = [INDICE_DIAS[d] for d in dias_validos]
        datas = todas_as_datas[todas_as_datas.dayofweek.isin(indices)]
        registros.append({
            "CURSO": aloc.get("CURSO"),
            "CODIGO": aloc.get("CODIGO"),
            "SALA": sala,
            "DISCIPLINA": aloc.get("DISCIPLINA"),
            "TURMA": aloc.get("TURMA"),
            "DIAS": ",".join(dias_validos),
            "HORARIO_INICIO": inicio_t.strftime("%H:%M") if inicio_t else None,
            "HORARIO_FINAL": fim_t.strftime("%H:%M") if fim_t else None,
            "HORARIOS_RAW": aloc.get("HORARIO") or aloc.get("HORÁRIO") or "",
            "ALUNOS": aloc.get("ALUNOS") or 0,
            "PROFESSOR": aloc.get("PROFESSOR"),
            "CAPACIDADE": next((s["CAPACIDADE"] for s in salas_ct if s["NOME"] == sala), None),
            "DATAS": datas,
            "DESCRICAO": descricao,
            "TIPO": "DISCIPLINA"
        })

        sala_obj = next((s for s in salas_ct if s["NOME"] == sala), None)
        if sala_obj:
            for d in dias_validos:
                if inicio_t and fim_t:
                    sala_obj["HORARIOS_OCUPADOS_SEMANA"][d].append((
                        inicio_t.strftime("%H:%M"), fim_t.strftime("%H:%M"), descricao
                    ))
                    sala_obj["HORARIOS_OCUPADOS"].add(f"{inicio_t.strftime('%H:%M')} - {fim_t.strftime('%H:%M')}")
                else:
                    raw = str(aloc.get("HORARIO") or "")
                    blocos = [b.strip() for b in raw.split(",") if b.strip()]
                    for bloco in blocos:
                        try:
                            parts = bloco.split()
                            dia = parts[0].upper()
                            horas = parts[1]
                            h1, h2 = horas.split("-")
                            if dia in DIAS_SEMANA:
                                sala_obj["HORARIOS_OCUPADOS_SEMANA"][dia].append((h1, h2, descricao))
                                sala_obj["HORARIOS_OCUPADOS"].add(f"{h1} - {h2}")
                        except Exception:
                            continue
    return pd.DataFrame(registros)

# -----------------------  Cria workbook por sala  -----------------------
def criar_workbook_horario_sala(sala_obj):
    horas_minutos = []
    for h in range(7, 22):
        horas_minutos.append(f"{h:02d}:00 - {h:02d}:30")
        horas_minutos.append(f"{h:02d}:30 - {h+1:02d}:00")

    wb = Workbook()
    ws = wb.active
    ws.title = sala_obj["NOME"][:31]

    dias = DIAS_SEMANA
    info_sala = f"Centro de Tecnologia \n {sala_obj['NOME']} - Período Letivo 2026.1 ({sala_obj['CAPACIDADE']})"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dias)+1)
    cell_info = ws.cell(row=1, column=1, value=info_sala)
    cell_info.font = Font(bold=True, size=12)
    cell_info.alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row=2, column=1, value="Horário").font = Font(bold=True)
    for col, dia in enumerate(dias, start=2):
        ws.cell(row=2, column=col, value=dia).font = Font(bold=True)

    for row, hora in enumerate(horas_minutos, start=3):
        ws.cell(row=row, column=1, value=hora)

    # ---------- preenche disciplinas + reservas ----------
    for col, dia in enumerate(dias, start=2):
        ocupados = sala_obj["HORARIOS_OCUPADOS_SEMANA"].get(dia, [])
        for inicio, fim, desc in ocupados:
            t_start = str_to_time(inicio)
            t_end = str_to_time(fim)
            if not t_start or not t_end:
                continue
            cur = dt.datetime.combine(dt.date.today(), t_start)
            fim_dt = dt.datetime.combine(dt.date.today(), t_end)

            # [CORREÇÃO 1] Verifica se é reserva manual
            is_reserva_manual = isinstance(desc, str) and desc.startswith("RESERVA_MANUAL")

            # [CORREÇÃO 2] Prepara o texto da célula e a fonte
            texto_celula = desc
            fonte_celula = Font(size=10)  # Fonte padrão para disciplinas

            if is_reserva_manual:
                # Extrai o nome do evento
                nome_evento = desc.replace("RESERVA_MANUAL - ", "").replace("RESERVA_MANUAL", "").strip()

                # [CORREÇÃO 3] Busca todas as datas desta mesma reserva pelo nome base
                datas_reserva = set()
                for r_data, r_ini, r_fim, r_desc in sala_obj["RESERVAS"]:
                    r_nome_base = r_desc.replace("RESERVA_MANUAL - ", "").replace("RESERVA_MANUAL", "").strip()
                    if r_ini == inicio and r_fim == fim and r_nome_base == nome_evento:
                        datas_reserva.add(r_data)

                # [CORREÇÃO 4] Formata o texto com as datas
                if len(datas_reserva) > 1:
                    data_ini_fmt = min(datas_reserva).strftime("%d/%m")
                    data_fim_fmt = max(datas_reserva).strftime("%d/%m")
                    texto_celula = f"RESERVA_MANUAL - {nome_evento} ({data_ini_fmt} a {data_fim_fmt})"
                elif len(datas_reserva) == 1:
                    data_ini_fmt = min(datas_reserva).strftime("%d/%m")
                    texto_celula = f"RESERVA_MANUAL - {nome_evento} ({data_ini_fmt})"
                else:
                    texto_celula = f"RESERVA_MANUAL - {nome_evento}"

                # [CORREÇÃO 5] Define fonte vermelha e negrito para reservas
                fonte_celula = Font(color="FF0000", bold=True, size=10)

            # Preenche todas as células do intervalo de tempo
            while cur < fim_dt:
                nxt = cur + dt.timedelta(minutes=30)
                label = f"{cur.time().strftime('%H:%M')} - {nxt.time().strftime('%H:%M')}"
                try:
                    row_idx = horas_minutos.index(label) + 3
                except ValueError:
                    cur = nxt
                    continue

                # [CORREÇÃO 6] Cria a célula UMA ÚNICA VEZ com o texto e aplica a fonte
                celula = ws.cell(row=row_idx, column=col, value=texto_celula)
                celula.font = fonte_celula

                cur = nxt

    # ---------- mescla células iguais ----------
    for col in range(2, len(dias) + 2):
        start_row = 3
        cur_val = ws.cell(row=3, column=col).value
        for row in range(3, len(horas_minutos) + 3):
            val = ws.cell(row=row, column=col).value
            if val != cur_val:
                if cur_val not in (None, "") and row - 1 >= start_row:
                    ws.merge_cells(start_row=start_row, start_column=col, end_row=row - 1, end_column=col)
                start_row = row
                cur_val = val
        if cur_val not in (None, "") and start_row <= len(horas_minutos) + 2:
            ws.merge_cells(start_row=start_row, start_column=col, end_row=len(horas_minutos) + 2, end_column=col)

    # ---------- estilo ----------
    thin = Side(style="thin")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=1, max_row=len(horas_minutos)+2, min_col=1, max_col=len(dias)+1):
        for cell in row:
            cell.border = borda
            cell.alignment = align
            if cell.font.size is None:
                cell.font = Font(size=10)

    for col in range(1, len(dias)+2):
        ws.column_dimensions[get_column_letter(col)].width = 30

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# -----------------------  Interface Streamlit  -----------------------
def interface_interativa(salas_ct, df_processado):
    st.header("🎯 Solicitação de Sala")

    evento = st.text_input("Digite o nome do evento:")
    nome = st.text_input("Digite seu nome:")
    email_cliente = st.text_input("Digite seu email:")
    capacity = st.number_input("Capacidade:", min_value=0, value=0)
    remetente = "solicitacaosalasct@gmail.com"
    destinatario = "reservasalact@ufc.br"
    assunto = "Teste - Solicitaçao de Salas"

    senha = "rmqz ohnf oppx zpwo"

    blocos = sorted({s["NOME"][:3] for s in salas_ct if s["NOME"]})
    bloco_sel = st.selectbox("Selecione o bloco:", blocos)
    salas_filt = [s["NOME"] for s in salas_ct if s["NOME"].startswith(bloco_sel)]
    sala_escolhida = st.selectbox("Selecione a sala:", salas_filt)

    col1, col2 = st.columns(2)
    with col1:
        data_ini = st.date_input("Data inicial:", key="dt_ini")
    with col2:
        usa_fim = st.selectbox("Data final (opcional):", ["NÃO", "SIM"], key="sn_fim")

    dias_evento = None
    if usa_fim == "SIM":
        data_fim = st.date_input("Data final:", key="dt_fim")
        dias_evento = st.multiselect("Dias da semana que o evento ocorrerá:", DIAS_SEMANA, default=["SEGUNDA"])
    else:
        data_fim = None

    h_ini = st.time_input("Horário de início:", key="h_ini")
    h_fim = st.time_input("Horário de término:", key="h_fim")

    sala_info = next((s for s in salas_ct if s["NOME"] == sala_escolhida), None)
    if sala_info is None:
        st.error("Sala não encontrada.")
        return

    st.subheader("Horários ocupados (por dia)")
    for dia in DIAS_SEMANA:
        ocu = sala_info["HORARIOS_OCUPADOS_SEMANA"].get(dia, [])
        st.write(f"**{dia}**: " + (", ".join([f"{a}-{b} ({c})" for a, b, c in ocu]) if ocu else "Nenhum"))

    # [CORREÇÃO 7] Inicializa session_state para armazenar reservas
    if 'reservas_lista' not in st.session_state:
        st.session_state.reservas_lista = []

    if 'df_completo' not in st.session_state:
        st.session_state.df_completo = df_processado.copy()

    if st.button("📅 Solicitar Sala", key="btn_solicitar"):
        inicio_str = h_ini.strftime("%H:%M")
        fim_str = h_fim.strftime("%H:%M")
        mapping = {'MONDAY': 'SEGUNDA', 'TUESDAY': 'TERÇA', 'WEDNESDAY': 'QUARTA',
                   'THURSDAY': 'QUINTA', 'FRIDAY': 'SEXTA', 'SATURDAY': 'SÁBADO', 'SUNDAY': 'DOMINGO'}

        # Prepara as datas a verificar
        if usa_fim == "SIM" and data_fim and dias_evento:
            datas_a_verificar = pd.date_range(data_ini, data_fim, freq='D') \
                                .to_series() \
                                .map(lambda d: mapping.get(d.strftime("%A").upper(),
                                                           d.strftime("%A").upper())) \
                                .isin(dias_evento)
            datas_a_verificar = pd.date_range(data_ini, data_fim, freq='D')[datas_a_verificar]
        else:
            datas_a_verificar = [data_ini]

        # [CORREÇÃO 8] Verifica conflitos - capacidade FORA do loop
        conflitos = []

        # Verifica capacidade apenas uma vez
        if capacity > sala_info["CAPACIDADE"]:
            ociosidade = (-1) * (sala_info["CAPACIDADE"] - capacity)
            conflitos.append({"tipo": "OCIOSIDADE", "valor": ociosidade})

        # Verifica conflitos de horário
        for data in datas_a_verificar:
            dia_port = mapping.get(data.strftime("%A").upper(), data.strftime("%A").upper())
            for a, b, desc in sala_info["HORARIOS_OCUPADOS_SEMANA"].get(dia_port, []):
                if intervals_overlap(a, b, inicio_str, fim_str):
                    conflitos.append({
                        "tipo": "HORARIO", 
                        "data": data.strftime("%d/%m"), 
                        "inicio": a, 
                        "fim": b, 
                        "desc": desc
                    })

        if conflitos:
            # [CORREÇÃO 9] Verifica tipo de conflito corretamente
            tem_ociosidade = any(c.get("tipo") == "OCIOSIDADE" for c in conflitos)
            tem_horario = any(c.get("tipo") == "HORARIO" for c in conflitos)

            if tem_ociosidade:
                ociosidade_val = next(c["valor"] for c in conflitos if c.get("tipo") == "OCIOSIDADE")
                st.error(f"❌ Conflito de Ociosidade: Capacidade excedida em {abs(ociosidade_val)} alunos (Capacidade da sala: {sala_info['CAPACIDADE']}, Solicitado: {capacity})")

            if tem_horario:
                msg_horarios = "\n".join([
                    f"• {c['data']}: {c['inicio']}-{c['fim']} ({c['desc']})" 
                    for c in conflitos if c.get("tipo") == "HORARIO"
                ])
                st.error(f"❌ Conflitos de horário encontrados:\n{msg_horarios}")

        else:
            # [CORREÇÃO 10] Prepara descrição consistente
            nome_evento = evento.strip() if evento and str(evento).strip() else "Evento Manual"
            desc = f"RESERVA_MANUAL - {nome_evento}"

            # Adiciona à sala em memória
            datas_list = []
            for data in datas_a_verificar:
                dia_port = mapping.get(data.strftime("%A").upper(), data.strftime("%A").upper())
                sala_info["RESERVAS"].append((data, inicio_str, fim_str, desc))
                sala_info["HORARIOS_OCUPADOS_SEMANA"].setdefault(dia_port, []).append(
                    (inicio_str, fim_str, desc))
                sala_info["HORARIOS_OCUPADOS"].add(f"{inicio_str} - {fim_str}")
                datas_list.append(data)

            # [CORREÇÃO 11] Cria registro da reserva para o DataFrame
            nova_reserva = {
                "CURSO": "RESERVA",
                "CODIGO": "MANUAL",
                "SALA": sala_escolhida,
                "DISCIPLINA": nome_evento,
                "TURMA": "N/A",
                "DIAS": ",".join([mapping.get(d.strftime("%A").upper(), d.strftime("%A").upper()) for d in datas_a_verificar]),
                "HORARIO_INICIO": inicio_str,
                "HORARIO_FINAL": fim_str,
                "ALUNOS": capacity,
                "PROFESSOR": nome,
                "CAPACIDADE": sala_info["CAPACIDADE"],
                "DATAS": datas_list,
                "DESCRICAO": desc,
                "TIPO": "RESERVA_MANUAL"
            }

            # Adiciona à lista de reservas no session_state
            st.session_state.reservas_lista.append(nova_reserva)

            # [CORREÇÃO 12] Atualiza o DataFrame completo com todas as reservas
            df_reservas = pd.DataFrame(st.session_state.reservas_lista)
            st.session_state.df_completo = pd.concat([df_processado, df_reservas], ignore_index=True)

            st.success(f"✅ Evento \"{nome_evento}\" registrado em {len(datas_a_verificar)} dia(s) na sala {sala_escolhida}.")
            st.info("💾 As reservas serão incluídas no download de 'dados_disciplinas.xlsx'")

    st.divider()

    # Botão de download do Excel da sala específica
    if st.download_button("📥 Baixar Excel (Sala)",
                          data=criar_workbook_horario_sala(sala_info),
                          file_name=f"horario_{sala_escolhida}.xlsx",
                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
        pass

    st.divider()
    
    # ==================== SEÇÃO DE SALVAMENTO NO GITHUB ====================
    st.subheader("💾 Salvar no Repositório GitHub")

    col1, col2 = st.columns(2)

    with col1:
        if st.button("☁️ Salvar no GitHub (substituir Resultados_Gerais.xlsx)", type="primary"):
            with st.spinner("Fazendo commit no GitHub..."):
                sucesso = commit_dados_disciplinas(st.session_state.df_completo)
                if sucesso:
                    st.success("✅ Arquivo atualizado no GitHub com sucesso!")
                    st.balloons()
                    st.info("🔄 A próxima vez que o app carregar, já terá os dados atualizados.")
                else:
                    st.error("❌ Falha ao salvar no GitHub. Verifique o token e as permissões.")

    with col2:
        # Mantém download como backup
        buf_df = BytesIO()
        st.session_state.df_completo.to_excel(buf_df, index=False, engine='openpyxl')
        buf_df.seek(0)
        st.download_button(
            "📥 Baixar cópia local (backup)", 
            data=buf_df,
            file_name="dados_disciplinas_backup.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # Mostra preview das reservas manuais
    if len(st.session_state.reservas_lista) > 0:
        st.divider()
        st.subheader("📋 Reservas Manuais Registradas nesta Sessão")
        df_preview = pd.DataFrame(st.session_state.reservas_lista)
        st.dataframe(df_preview[["SALA", "DISCIPLINA", "DIAS", "HORARIO_INICIO", "HORARIO_FINAL", "ALUNOS", "PROFESSOR"]], 
                    use_container_width=True)

# -----------------------  Main  -----------------------
def main():
    st.title("🏫 Sistema de Alocação de Salas – CT")
    with st.spinner("Carregando dados..."):
        arquivo_sala = pd.read_excel(CAMINHO_SALAS)
        arquivo_disciplina = pd.read_excel(CAMINHO_DISCIPLINAS)
        df_salas, df_turmas = arquivo_sala, arquivo_disciplina
        salas_ct = criar_lista_salas(df_salas)
        todas_as_datas = gerar_datas(df_turmas)
        df_dados = processar_alocacoes(df_turmas, todas_as_datas, salas_ct)
    st.success("✅ Dados carregados e processados com sucesso!")
    st.divider()
    interface_interativa(salas_ct, df_dados)

if __name__ == "__main__":
    main()
